from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import shutil
import re
import os

def get_single_zip_file(directory_path):
    if not os.path.exists(directory_path):
        raise FileNotFoundError(f"The directory '{directory_path}' does not exist.")
    zip_files = [f for f in os.listdir(directory_path) if f.endswith('.zip')]
    if len(zip_files) == 0:
        raise ValueError("No zip files found in the directory.")
    elif len(zip_files) > 1:
        raise ValueError("Multiple zip files found in the directory.")
    return zip_files[0].split(".")[0]

def set_info_sheet_column_widths(filename):
    wb = load_workbook(filename)
    if 'Info' in wb.sheetnames:
        ws = wb['Info']
        ws.column_dimensions['B'].width = 60
        for column in ['C', 'D', 'E']:
            ws.column_dimensions[column].width = 25
        wb.save(filename)
    else:
        print("No 'Info' sheet found in the workbook.")

    for sheet in wb.sheetnames:
        if sheet != 'Info':
            ws = wb['Risk Register']
            for column in ['P', 'Q']:
                ws.column_dimensions[column].width = 25
            wb.save(filename)
        else:
            print("No 'Risk Register' sheet found in the workbook.")

def adjust_column_widths(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max(max_length + 2, 8)
    wb.save(filename)

def add_border_to_row(sheet, row_number, start_col, end_col):
    light_blue_border = Border(
        left=Side(border_style="thin", color="ADD8E6"),
        right=Side(border_style="thin", color="ADD8E6"),
        top=Side(border_style="thin", color="ADD8E6"),
        bottom=Side(border_style="thin", color="ADD8E6"),
    )
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row_number, column=col)
        cell.border = light_blue_border

def increaseWidth(filename):
    try:
        workbook = load_workbook(filename)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for col_index in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_index)
                worksheet.column_dimensions[col_letter].width = 20
            for row_index in range(1, worksheet.max_row + 1):
                worksheet.row_dimensions[row_index].height = 18
        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
        )
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=False,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )
        light_blue_border = Border(
        left=Side(border_style="thin", color="ADD8E6"),
        right=Side(border_style="thin", color="ADD8E6"),
        top=Side(border_style="thin", color="ADD8E6"),
        bottom=Side(border_style="thin", color="ADD8E6"),
        )
        target_cell.border = light_blue_border


def getText(arr):
    r = []
    for i in arr: r.append(i.text.split("\n"))
    return r

def got(arr):
    r = []
    for i in arr: r.append(i.text.split("\n")[0])
    return r

def getCategories(id):
    with open(f"./files/{id}/file.xml", "r+") as f:
        soup = BeautifulSoup(f.read(), "xml")
    return got(soup.find_all("Category"))

def getData(id):
    with open(f"./files/{id}/file.xml", "r+") as f:
        data=readXML(f.read())
    return data


def readXML(content):
    soup = BeautifulSoup(content, "xml")
    arr = []
    primaries = getText(soup.find_all("Primary"))
    categories = getText(soup.find_all("Category"))
    for primary in primaries:
        app = {}
        app['filename'] = primary[1]
        app['filepath'] = primary[2]
        app['linenumber'] = primary[3]
        app['category'] = categories[(primaries.index(primary))]
        arr.append(app)
    return arr

def createExcel(id, data, system, tester, date):
    filepath = f'./files/{id}/file.xlsx'
    file_path = 'Tracker Template.xlsx'
    file_path2 = 'Category Template.xlsx'
    categories = getCategories(id)
    shutil.copy(file_path, filepath)
    for category in set(categories):
        workbook = load_workbook(filepath)
        sheet = workbook["Risk Register"]
        new_entry = [
            sheet.max_row-1,  # S/N
            system , # System
            "Code Review",  # Area of Review
            "",  # Overall Risk Rating
            "",  # Impact Rating
            "",  # Likelihood Rating
            "",  # CVSS Score
            "",  # CVSS Vector
            f'Refer to worksheet: "{category}"',  # Affected File
            f'Refer to worksheet: "{category}"',  # Affected Line
            categories.count(category),  # Count of affected file
            f'Refer to worksheet: "{category}"',  # Screenshot
            category,  # Issue Title
            "",  # Observations
            "",  # Implications
            "",  # Recommendations
            "",  # Management Comments
            date,  # Date Raised
            tester,  # DT Tester
            "",  # Client Owned
            "Open",  # Status
            "",  # Date Follow-up
            "",  # DT Tester
            "",  # Post Review Observations
            "",  # Post Review Screenshots
            ""  # Client Remarks
        ]
        sheet.append(new_entry)
        new_row_number = sheet.max_row
        add_border_to_row(sheet, new_row_number, 1, len(new_entry))
        print(new_row_number)
        workbook.save(filepath)
        
    workbook1 = load_workbook(file_path2)
    source_sheet = workbook1["Category"]
    categories = getCategories(id=id)
    for category in set(categories):
        counter = 1
        workbook2 = load_workbook(filepath)
        new_sheet_name = re.sub(INVALID_TITLE_REGEX, '_', category)[:30]
        if new_sheet_name not in workbook2.sheetnames:
            target_sheet = workbook2.create_sheet(new_sheet_name)
        else:
            target_sheet = workbook2[new_sheet_name]
        for row_index, row in enumerate(source_sheet.iter_rows(), start=1):
            for col_index, cell in enumerate(row, start=1):
                target_cell = target_sheet.cell(row=row_index, column=col_index)
                copy_cell(cell, target_cell)
        workbook2.save(filepath)
        workbook2 = load_workbook(filepath)
        for datapoint in data:
            if datapoint['category'][0] == category:
                sheet = workbook2[new_sheet_name]
                new_entry = [
                        sheet.max_row-1,  # S/N
                        system,  # System
                        "Code Review",  # Area of Review
                        "",  # Overall Risk Rating
                        "",  # Impact Rating
                        "",  # Likelihood Rating
                        "",  # CVSS Score
                        "",  # CVSS Vector
                        datapoint["filepath"],  # Affected File
                        f"{datapoint['linenumber']}",  # Affected Line
                        "",  # Screenshot
                        category,  # Issue Title
                        "",  # Observations
                        "",  # Implications
                        "",  # Recommendations
                        "",  # Management Comments
                        date,  # Date Raised
                        tester,  # DT Tester
                        "",  # Client Owned
                        "Open",  # Status
                        "",  # Date Raised
                        "",  # DT Tester
                        "",  # Post Review Observations
                        "",  # Post Review Screenshots
                        ""  # Client Remarks
                    ]
                sheet.append(new_entry)
                new_row_number = sheet.max_row
                add_border_to_row(sheet, new_row_number, 1, len(new_entry))
                workbook2.save(filepath)
                counter+=1
    
    adjust_column_widths(filepath)
    set_info_sheet_column_widths(filepath)